import sys
import json
from openpyxl import Workbook
from totra.exceptions import UnknownFormat


def format_activities(json_content, format='json', format_params = {}):
    if format == 'json':
        output = json.dumps(json_content, indent=4)
    elif format == 'excel':
        output = convert_json_activities_to_excel(json_content, format_params)
    else:
        raise UnknownFormat
    return output


def save_output(output, output_filename):
    if not output_filename:
        return
    try:
        output.save(output_filename)
    except AttributeError:
        if output_filename == 'stdout':
            sys.stdout.write(str(output))
        else:
            with open(output_filename, 'w') as outfile:
                outfile.write(str(output))


def convert_json_activities_to_excel(json_activities, format_params):
    duration_types = {'hours': 3600, 'minutes': 60, 'seconds': 1}

    duration = format_params.get('duration', 'minutes')
    duration_precision = format_params.get('duration_precision', 2)
    duration_div = duration_types[duration]

    wb = Workbook()
    ws = wb.active
    # header
    row = 1
    col = 1
    header = ['Project name', 'Description', 'Worker', 'Start time', 'End time', 'Duration']
    for header_elem in header:
        ws.cell(column=col, row=row, value=header_elem)
        col += 1
    # values
    row = 2
    for activity in json_activities:
        col = 1
        values = [activity['project']['name'], activity['description'],
                  activity['worker']['name'], activity['start_time'],
                  activity['end_time'], round(activity['seconds'] / duration_div, duration_precision)]
        for value in values:
            ws.cell(column=col, row=row, value=value)
            col += 1
        row += 1
    return wb
